from django.http import HttpResponse
from django.shortcuts import render
import xlsxwriter
from openpyxl import load_workbook

workbook = xlsxwriter.Workbook("1.xlsx")
worksheet = workbook.add_worksheet("1sheet")
workbook.close()



def index(request):
    return render(request, 'index.html')

    # return HttpResponse("Home")






def analyze(request):
    #Get the text
    first_name = request.GET.get('fname', 'default')
    second_name = request.GET.get('lname','default')
    workbook=load_workbook(filename="1.xlsx")
    sheet=workbook.active
    sheet["A1"]="1"
    sheet["B1"]=first_name
    sheet["C1"]=second_name
    sheet["A2"]="2"
    sheet["B2"]=first_name
    sheet["C2"]=second_name
    
    print(first_name)
    print(second_name)
    
    workbook.save(filename="1.xlsx")
    
    return HttpResponse('''You are good to go''')


